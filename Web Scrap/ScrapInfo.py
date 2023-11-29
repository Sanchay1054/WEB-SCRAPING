import bs4
import requests
import spacy
import json
from spacy.training.example import Example
import spacy
import warnings
from collections import Counter
import openpyxl
import threading
import tkinter as tk

company_name = "tcpl"
designation = "ceo"


nlp = spacy.blank("en")
ner = nlp.add_pipe("ner")
#ner.add_label("name")
#ner.add_label("designation")
ner.add_label("inlinked")
ner.add_label("google")
ner.add_label("website")
ner.add_label("question")
ner.add_label("symbol")
ner.add_label("company")
ner.add_label("phone_number")
#ner = nlp.add_pipe('name')

#ner.add_label("title")
#ner.add_label("subject")
#ner.add_label("company")
#ner.add_label("university")
#warnings.filterwarnings("ignore",message=".*'spacy.training.offsets_to_biluo_tags'.*")
train_data = [
    ("ram is a good boy",{"entities":[(0,3,"name"),(9,13,"designation")]}),
    ("sunil d'souze is ceo of tcpl ",{"entities":[(0,5,"name"),(6,13,"name"),(17,20,"designation"),(24,28,"company"),(7,7,"symbol")]}),
    ("John and Mary are cofounders of ABC Ltd", {"entities": [(0, 4, "name"), (9, 13, "name")]}),
    ("Professor Johnson teaches physics at University XYZ", {"entities": [(10, 17, "name"),(37,47,"company")]}),
    (". , = > : ( ) { } [ ] * & - | = ",{"entities":[(0,1,"symbol"),(2,3,"symbol"),(4,5,"symbol"),(6,7,"symbol"),(8,9,"symbol"),(10,11,"symbol"),(12,13,"symbol"),(14,15,"symbol"),(16,17,"symbol"),(18,19,"symbol"),(20,21,"symbol"),(22,23,"symbol"),(24,25,"symbol"),(26,27,"symbol"),(28,29,"symbol"),(30,31,"symbol")]}),
    ("httpsin.linkedin/in ",{"entities":[(0,19,"inlinked")]}),
    ("httpsin.linkedin/in ",{"entities":[(0,19,"inlinked")]}),
    ("linkedin.in ",{"entities":[(0,11,"inlinked")]}),
    ("linkedin ",{"entities":[(0,8,"inlinked")]}),
    ("ram from chennai is md ",{"entities":[(0,3,"name"),(20,22,"designation")]}),
    ("chandrasekaran is a ceo of tcpl theni ",{"entities":[(0,14,"name"),(20,23,"designation")]}),
    ("sunil from chennai is Souza playing ",{"entities":[(0,5,"name")]}),
    ("sameer is a khetarpal enjoying ceo of jubilant ",{"entities":[(0,6,"name"),(12,21,"name"),(31,34,"designation")]}),
    ("venu of a Nair good chilling ceo from chennai is md ",{"entities":[(0,4,"name"),(29,32,"designation")]}),
    ("milind second Pant from madurai is ceo ",{"entities":[(0,6,"name"),(7,13,"name"),(14,18,"name"),(39,42,"designation")]}),
    ("david from India is playing chess ceo ",{"entities":[(0,5,"name"),(34,37,"designation")]}),
    ("akash Poddar from chennai is blowing ceo ",{"entities":[(0,5,"name"),(6,12,"name"),(37,40,"designation")]}),
    ("sanjeev from gujarat is meeting kumar ceo ",{"entities":[(0,7,"name"),(38,41,"designation")]}),
    ("mahesh from kashmir is gupta hr ",{"entities":[(0,6,"name"),(29,31,"designation")]}),
    ("rohit from chennai is md ",{"entities":[(0,5,"name"),(22,24,"designation")]}),
    ("rajiv from chennai is md ",{"entities":[(0,5,"name"),(22,24,"designation")]}),
    ("robert from chennai is md ",{"entities":[(0,6,"name"),(23,25,"designation")]}),
    ("sanjiv from chennai is md ",{"entities":[(0,6,"name"),(23,25,"designation")]}),
    ("yuki from chennai is Managing Director ",{"entities":[(0,4,"name"),(21,29,"designation"),(30,38,"designation")]}),
    ("rajeev from chennai is Vice president ",{"entities":[(0,6,"name"),(23,27,"designation"),(28,37,"designation")]}),
    ("sujith from chennai is Executive Director ",{"entities":[(0,6,"name"),(23,32,"designation"),(33,41,"designation")]}),
    ("sudir from chennai is sitapati human resourse",{"entities":[(0,5,"name"),(22,30,"name"),(37,36,"designation"),(37,45,"name")]}),
    ("+919275927484 ",{"entities":[(0,13,"phone_number")]}),
    ("0443273892888 ",{"entities":[(0,13,"phone_number")]}),
    ("(91)915556662 ",{"entities":[(0,13,"phone_number")]}),
    ("9374974729 ",{"entities":[(0,10,"phone_number")]}),
    ("8366836582 ",{"entities":[(0,10,"phone_number")]})
    ]

other_pipes  = [pipe for pipe in nlp.pipe_names if pipe!="name"]
example = []


with nlp.disable_pipes(*other_pipes):
    for text,annotations in train_data:
        #print(text)
        #print([(start_label,end_label,label) for start_label, end_label, label in annotations["entities"]])
        doc = nlp.make_doc(text)
        #span = [doc[start:end].char_span(start, end, label=label) for start, end, label in annotations["entities"]]
        #doc = spacy.training.offsets_to_biluo_tags(spacy.tokens.doc.Doc(nlp.vocab,words=text.split()),annotations["entities"])
        #train_data_example = (doc,annotations)
        examples = Example.from_dict(doc,annotations)
        #print(examples)
        example.append(examples)

AIwindow = tk.Tk()
text = tk.Label(AIwindow,text="Training Model Please Wait")
text.pack()


def start_train():
    nlp.to_disk("trained_data.spacy")
    nlp.begin_training()
    for epoch in range(25):
        for words in example:
            losses ={}
            nlp.update(example,drop=0,losses=losses)
            #print(words.reference.ents)
text_data = "Ram is ceo in tcpl d'souze"
#doc = nlp(text_data)
#print("Entities",[(ent.text, ent.label_) for ent in doc.ents],doc.ents)

#text.config(text="ScrapInfo")


#print("Entities",[(ent.text, ent.label_) for ent in doc_soup.ents],"\n",max(set(" ".join(text_soup.split(" "))),key=" ".join(text_soup.split(" ")).count))

def name():
    try:
        name_company = name_text.get()
        post = designation_text.get()
        url = "https://www.google.com/search?q= "+name_company+" "+post
        search_button.config(text="SEARCHING")
        page = requests.get(url)

        soup = bs4.BeautifulSoup(page.text,'html.parser')
        nlp_soup = {"name":[],"designation":[],"company":[],"symbol":[],"inlinked":[],"phone_number":[]}
        line_soup = soup.find("body").find_all("div")
        name = []

        text_soup = soup.find("body").text.lower().replace(",","").replace("-","").replace(".","").replace(">","").replace("<","").replace("·","")
        doc_soup = nlp(text_soup)

        for text in line_soup:
            line_doc = nlp(text.text.lower())
            for ent in line_doc.ents:
                if ent.label_!="inlinked":
                    nlp_soup[ent.label_].append(ent.text)
            #print("Entities",line_doc.text,[(ent.text, ent.label_) for ent in line_doc.ents])

        #print(line_soup.text)
        name = Counter(nlp_soup["name"]).most_common(3)

        url = url+" linkedin"
        page = requests.get(url)
        soup = bs4.BeautifulSoup(page.text,'html.parser')

        link_soup = soup.find("body").find_all("a")
        #print(name[0][0])

        for text in link_soup:
            link_href = text["href"]
            link_name = text.text.lower()
            #print(link_href+" "+link_name)
            '''link_doc = nlp(text["href"])
            name_doc = nlp(text.text)
            print(text.text)
            for ent,name_en in zip(link_doc.ents,name_doc.ents):
                #print(text)'''
            if "linkedin" in link_href and name[0][0] in link_name:
                nlp_soup["inlinked"].append(text["href"])
                #print("Entities",(ent.text,ent.label_))



        url = "https://www.google.com/search?q= "+name_company+" office phone numbe"

        page = requests.get(url)
        soup = bs4.BeautifulSoup(page.text,'html.parser')

        phone_number_soup = soup.find("body").find_all("div")
        nlp_soup["phone_number"] =[]

        for text in phone_number_soup:
            phone_number = text.text.split()
            for phone_text in phone_number:
                phone_number_text = nlp(phone_text)
                for ent in phone_number_text.ents:
                    nlp_soup[ent.label_].append(phone_text)
                    #print(text.text)

        phone_number = Counter(nlp_soup["phone_number"]).most_common(10)
        try:
            test = name[0][0]
        except:
            name[0][0] = ""
        try:
            test = name[1][0]
        except:
            name[1][0] = ""
        try:
            test = name[2][0]
        except:
            name[2][0] = ""
        try:
            linkedin_name = nlp_soup["inlinked"][0].split("=")[1].split("/")[4]
        except:
            linkedin_name = ""
            #nlp_soup["inlinked"][0].split("=")[0] = "/////"
            #nlp_soup["inlinked"][0].split("=")[1] = "/////"
        try:
            linkedin_domain = nlp_soup["inlinked"][0].split("=")[1]
        except:
            linkedin_domain = ""
            #nlp_soup["inlinked"][0].split("=")[0] = "/////"
            #nlp_soup["inlinked"][0].split("=")[1] = "/////"
        file_name = openpyxl.load_workbook("ScrapExcel.xlsx")
        work_sheet = file_name.active
        row_work = 1
        coloumn_work = 1
        row_name = work_sheet.cell(row=1,column=1).value
        while work_sheet.cell(row=row_work,column=1).value!=None:
            row_work+=1
            row_name = work_sheet.cell(row=row_work,column=coloumn_work).value
            #print(row_name)
        work_sheet.cell(row=row_work,column=coloumn_work).value = name[0][0]+" "+name[1][0]+" "+name[2][0]+str("".join([words for text in linkedin_name.split("-") for words in text if words not in ["1","2","3","4","5","6","7","8","9","0"]]))
        work_sheet.cell(row=row_work,column=2).value = str(linkedin_domain)
        work_sheet.cell(row=row_work,column=3).value = " or ".join([phone_numbers[0] for phone_numbers in phone_number if len(phone_numbers[0])>=4 and phone_numbers[0] not in "abcdefghijklmnopqrstuvwxyz"])
        file_name.save("ScrapExcel.xlsx")
        #print(work_sheet.cell(row=1,column=1).value!=None,row_work)
        name_box.config(text=str("NAME: "+name[0][0]+" "+name[1][0]+" "+name[2][0]+"\nData From LinkedIn: "+str("".join([words for text in linkedin_name.split("-") for words in text if words not in ["1","2","3","4","5","6","7","8","9","0"]]))))
        linkedin_box.config(text="LINKEDIN URL: "+str(linkedin_domain))
        phone_number_box.config(text="PHONE NUMBER: "+" or ".join([phone_numbers[0] for phone_numbers in phone_number if len(phone_numbers[0])>=4 and phone_numbers[0] not in "abcdefghijklmnopqrstuvwxyz"]))
        search_button.config(text="SEARCH")
        
        
        '''for names in name:
            print(names[0])
        for phone_numbers in phone_number:
            if len(phone_numbers[0])>=4 and phone_numbers[0] not in "abcdefghifklmnopqrstuvwxyz":
                print(phone_numbers[0])
        print(nlp_soup["inlinked"][0])'''
    except:
        search_button.config(text="SEARCH")
        name_box.config(text="Check your Internet Connection")
        linkedin_box.config(text="Check your internet connection")
        phone_number_box.config(text="Check your internet connection")


#if page.status_code==200:
    #print("page opened")
            
#print(nlp_soup)

def button():
    start_train()
    global name_text,name_text_box,designation_text,designation_text_box,search_button,name_box,linkedin_box,phone_number_box
    name_text = tk.StringVar()
    name_text_box = tk.Entry(AIwindow,textvariable=name_text)
    name_text_box.pack()

    designation_text = tk.StringVar()
    designation_text_box = tk.Entry(AIwindow,textvariable=designation_text)
    designation_text_box.pack()
    text.config(text="ScrapInfo\n\nType Company name in first box \n\nDesignation in second box (ceo)")
    search_button = tk.Button(AIwindow,text="SEARCH",command=lambda : threading.Thread(target=name).start())
    search_button.pack()

    name_box = tk.Label(AIwindow,text="NAME")
    name_box.pack()

    linkedin_box = tk.Label(AIwindow,text="LINKEDIN")
    linkedin_box.pack()

    phone_number_box = tk.Label(AIwindow,text="PHONE NUMBER")
    phone_number_box.pack()



#print(soup.find("body"))
threading.Thread(target=button).start()
AIwindow.mainloop()
